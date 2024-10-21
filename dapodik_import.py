import subprocess
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from dotenv import load_dotenv
from colorama import Fore
import colorama
import hashlib
import json
import os

colorama.init(autoreset=True)

class DapodikImport:
    def __init__(self):
        load_dotenv(override=True)
        self.column_nis = os.environ.get("column_nis",False)
        self.column_name = os.environ.get("column_name",False)
        self.low_dapodik_import = LowDapodikImport()
    
    def lets_import(self):
        self.low_dapodik_import.set_excel_path("./import_dapodik/excel/dapodik.xlsx")
        if not self.column_name or not self.column_nis:
            print(f"{Fore.RED}>> Select column for import")
            print(f"{Fore.RED}>> [Help] (Replace X & Y) add column_nis='X' and column_name='Y' in your .env file")
            exit(1)
        self.low_dapodik_import.set_column_config(column_nis=self.column_nis, column_name=self.column_name)
        self.low_dapodik_import.process_and_dump_to("./db/main.json")

class LowDapodikImport:
    def __init__(self):
        self.__import_folder = "import_dapodik"
        self.__excel_path = ""
        self.__column_nis = 0
        self.__column_name = 0
        self.__generated_hash = f"./{self.__import_folder}/prev_sha256.json"
        self.__excel_folder = f"./{self.__import_folder}/excel"
        self.__rclone_invoke = ["rclone", "sync", "-v", "--retries", "10"]
        self.__var_sync_foto = ""
        self.__var_sync_excel = ""
        self.__init_default_sync_foto("anshar:QRAnsharFoto")
        self.__init_default_sync_excel("anshar:QRAnsharDapodik")

    def set_excel_path(self, file_path:str):
        if not os.path.exists(self.__import_folder):
            os.mkdir(self.__import_folder)
        if not os.path.exists(self.__excel_folder):
            os.mkdir(self.__excel_folder)
        if os.path.exists(file_path):
            self.__excel_path = file_path
        else:
            print(f"{Fore.RED}File not found ...")
            exit(1)

    def set_column_config(self, column_nis:str, column_name:str):
        self.__column_nis = column_index_from_string(column_nis) - 1
        self.__column_name = column_index_from_string(column_name) - 1
    
    def process_and_dump_to(self, output_path:str):
        print(f"{Fore.CYAN}>> Checking for new file from google drive ...")
        self.__sync_foto()
        self.__sync_excel()
        if self.__is_changed_file():
            load = self.__check_if_output_already_exist(output_path = output_path)
            loaded = self.__load_from_xlsx(load_existing = load)
            with open(output_path,"w") as f:
                json.dump(loaded,f)
            print(f"{Fore.CYAN}>> Imported ...")
        else:
            # Up To Date
            print(f"{Fore.CYAN}>> No changes from import dapodik yet ...")

    def modify_sync_foto(self, download_to:str, remote:str, remote_folder:str = "QRAnsharFoto"):
        self.__var_sync_foto = [f"{remote}:{remote_folder}",f"{download_to}"]

    def modify_sync_excel(self, download_to:str, remote:str, remote_folder:str = "QRAnsharExcel"):
        self.__var_sync_excel = [f"{remote}:{remote_folder}",f"{download_to}"]

    def __init_default_sync_foto(self, remote_and_remote_folder):
        self.__var_sync_foto = [f"{remote_and_remote_folder}",f"{os.path.join(os.getcwd(),'static/assets/student-pictures')}"]

    def __init_default_sync_excel(self, remote_and_remote_folder):
        self.__var_sync_excel = [f"{remote_and_remote_folder}",f"{os.path.join(os.getcwd(),'import_dapodik/excel')}"]

    def __check_if_output_already_exist(self, output_path):
        if not os.path.exists(output_path):
            return {}
        with open(output_path,"r") as f:
            return json.load(f)
    
    def __load_from_xlsx(self, load_existing):
        workbook = load_workbook(self.__excel_path)
        worksheet = workbook.active
        load = load_existing
        rows_skipped = 0
        skipped_index = []
        for row_num,row in enumerate(worksheet.iter_rows(values_only=True)):
            try:
                if row[0]:
                    nomer_induk_siswa = int(row[self.__column_nis])
                    nama_siswa = str(row[self.__column_name])
                    load[f"stu-id-{nomer_induk_siswa}"] = {"nama":nama_siswa}
                else:
                    rows_skipped+=1
                    skipped_index.append(row_num)
            except Exception as e:
                print(e)
                print(f"{Fore.RED}>> Error wrong data type [dapodik_import]")
                print(f"{Fore.RED}>> Check column name .env or possible defect dapodik")
                workbook.close()
                exit(1)
        if rows_skipped:
            print(f"{Fore.YELLOW}>> Encountered none values, skipped rows : {rows_skipped} at indices : {skipped_index}")
        workbook.close()
        return load

    def __is_changed_file(self):
        # 500Kb
        BUF_SIZE = 524288
        previous_hash = self.__load_prev_hash().get("sha256").lower()
        current_hash = hashlib.sha256()
        with open(self.__excel_path, 'rb') as f:
            while True:
                data = f.read(BUF_SIZE)
                if not data:
                    break
                current_hash.update(data)
        current_hash = current_hash.hexdigest()
        if previous_hash != current_hash:
            with open(self.__generated_hash,"w") as f:
                json.dump({"sha256":current_hash},f)
            return True
        else:
            return False

    def __load_prev_hash(self):
        if os.path.exists(self.__generated_hash):
            with open(self.__generated_hash,"r") as f:
                return json.load(f)
        return {"sha256":"init"}

    def __sync_foto(self):
        sync_foto = self.__rclone_invoke + self.__var_sync_foto
        return self.__run_command(sync_foto)

    def __sync_excel(self):
        sync_excel = self.__rclone_invoke + self.__var_sync_excel
        return self.__run_command(sync_excel, is_wait = True)

    def __run_command(self, command, is_wait = False):
        try:
            if is_wait:
                print(f"{Fore.YELLOW}>> Waiting for important sync to finish ...")
                subprocess.Popen(command).wait()
            else:
                print(f"{Fore.YELLOW}>> Sync started and not waited ...")
                subprocess.Popen(command)
        except Exception as e:
            print(f"{Fore.RED}>> Something went wrong in run command")
            print(f"{Fore.RED}>> Error: {e}")

if __name__ == "__main__":
    dpdk = DapodikImport()
    dpdk.lets_import()